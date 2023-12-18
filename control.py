from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QImage, QPixmap,QIcon
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtWidgets import QApplication
#視窗模組
import cv2
#電腦視覺模組
import numpy
#數學處理模組
import os
#系統模組
from ui import Ui_MainWindow
#導入UI
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
#excel模組

class MainWindow_controller(QtWidgets.QMainWindow): #主要視窗呼叫
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    def __init__(self):
        super().__init__() # in python3, super(Class, self).xxx = super().xxx
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_control()
        QMainWindow.setWindowTitle(self,"UwU")
        self.setWindowIcon(QIcon('cheese.ico'))

    #視窗以及標題設定
        

    def setup_control(self):
        self.ui.folderbutton.clicked.connect(self.open_folder)
        self.ui.targetbutton.clicked.connect(self.target_folder)
        self.ui.start.clicked.connect(self.execute)
    #對不同按鈕指定到不同function

    def open_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self,
                  "Open folder",
                  "./")                 # start path
        #print(folder_path)
        self.ui.folderpath.setText(folder_path)

    def target_folder(self):
        target_path = QFileDialog.getExistingDirectory(self,
                  "Open folder",
                  "./")                 # start path
        #print(target_path)
        self.ui.targetpath.setText(target_path)
    #開啟資料夾的function

    def execute(self):#分析的function
        self.ui.start.setDisabled(True)
        self.ui.start.setText("analysing")
        QApplication.processEvents()
        #畫面更新
        targetpath = self.ui.targetpath.text()
        os.chdir(targetpath)
        #目標定位
        wb = openpyxl.Workbook()
        sheet_name = []
        sheet_count = {}
        ffont = Font(name='Arial',size = 10)
        #格式變數

        #同analyze.py
        folderspath = self.ui.folderpath.text()
        position_x = int(self.ui.pos_x.text())
        position_y = int(self.ui.pos_y.text())
        width = int(self.ui.pos_w.text())
        hight = int(self.ui.pos_h.text())

        position_xb = int(self.ui.pos_x_b.text())
        position_yb = int(self.ui.pos_y_b.text())
        widthb = int(self.ui.pos_w_b.text())
        hightb = int(self.ui.pos_h_b.text())

        file_amount = 0
        file_count = {}
        #ori_path = os.path.dirname(os.path.realpath(__file__)) #程式所在位置 不可用
        os.chdir(folderspath)
        file_list = []
        first_point = ((position_x),(position_y))
        first_point_b = ((position_xb),(position_yb))
        second_point = (first_point[0]+width,first_point[1]+hight)
        second_point_b = (first_point_b[0]+widthb,first_point_b[1]+hightb)
        folder_list = os.listdir(folderspath)


        for name in folder_list:
            file_amount = file_amount + 1 
            file_list.append(name)

        for name in file_list:
            front_p = name.rfind('_')
            #dot_p = name.rfind('.')
            front = name[:front_p]
            #dot = name[dot_p:]
            if(front not in sheet_name):
                sheet_name.append(front)
                wb.create_sheet(front)
                sheet_count[front] = 3

        file_list.sort(key= lambda x:int(x[x.rfind('X')+1:x.rfind('.')]))

        


        for times in range(file_amount):
            self.ui.start.setText("analysing")
            QApplication.processEvents()
            img_name = file_list[times]
            #print(times)
            img_origin = cv2.imread(img_name,-1)
            img_frag_sample = img_origin[first_point[1]:second_point[1], first_point[0]:second_point[0]]#y0:y1, x0:x1
            img_frag_reference = img_origin[first_point_b[1]:second_point_b[1], first_point_b[0]:second_point_b[0]]#y0:y1, x0:x1
            siz_sample = img_frag_sample.shape
            #print(siz_sample)
            siz_reference = img_frag_reference.shape

            for yt in range(siz_sample[1]):
                summ =0
                for xt in range(siz_sample[0]):
                    summ +=img_frag_sample[xt][yt]
                summ = summ/siz_sample[0]
                self.ui.start.setText("file writing")
                img_sheet = img_name[:img_name.rfind('_')]
                shte = wb[img_sheet]
                shte.cell(yt+2+position_x,sheet_count[img_sheet]).value = summ
                #print(str(yt+2)+' '+str(sheet_count[img_sheet])+' '+str(summ))
                shte.cell(yt+2+position_x,sheet_count[img_sheet]).font = ffont
            shte.cell(1,sheet_count[img_sheet]).value = img_name[img_name.rfind('_')+1:img_name.rfind('.')]+"_sample"
            shte.cell(1,sheet_count[img_sheet]).font = ffont
            letter1 = shte.cell(1,sheet_count[img_sheet]).column
            letter2 = get_column_letter(letter1)
            shte.column_dimensions[letter2].width = 10.0
            sheet_count[img_sheet] = sheet_count[img_sheet]+1
                
            for yt in range(siz_reference[1]):
                summ = 0
                for xt in range(siz_reference[0]):
                    summ += img_frag_reference[xt][yt]
                summ = summ / siz_reference[0]
                self.ui.start.setText("file writing")
                QApplication.processEvents()
                img_sheet = img_name[:img_name.rfind('_')]
                shte = wb[img_sheet]
                shte.cell(yt+2+position_xb, sheet_count[img_sheet]).value = summ
                #print(str(yt+2)+' '+str(sheet_count[img_sheet])+' '+str(summ))
                shte.cell(yt+2+position_xb, sheet_count[img_sheet]).font = ffont
            shte.cell(1, sheet_count[img_sheet]).value = img_name[img_name.rfind('_')+1:img_name.rfind('.')]+"_reference"
            shte.cell(1, sheet_count[img_sheet]).font = ffont
            letter1 = shte.cell(1, sheet_count[img_sheet]).column
            letter2 = get_column_letter(letter1)
            shte.column_dimensions[letter2].width = 10.0
            sheet_count[img_sheet] = sheet_count[img_sheet]+1
            
            if(not(img_sheet in file_count)):
                file_count[img_sheet] = 1
            else:
                file_count[img_sheet] = file_count[img_sheet]+1

        excelname = folderspath[folderspath.rfind('/')+1:]
        os.chdir(targetpath)
        wb.save(excelname+'.xlsx')
        wb = openpyxl.load_workbook(excelname+'.xlsx')
        os.chdir(folderspath)
        for sh in range(len(sheet_count)):
            self.ui.start.setText("file creating")
            QApplication.processEvents()
            avsample = 0
            avreference = 0
            shte = wb.worksheets[sh+1]
            all = sheet_count[sheet_name[sh]]
            shte.cell(1, all+1).value = "Averaged"
            shte.cell(1, all+1).font = ffont
            shte.cell(1, all+2).value = "sample"
            shte.cell(1, all+2).font = ffont
            shte.cell(1, all+3).value = "reference"
            shte.cell(1, all+3).font = ffont
            shte.cell(1, 1).value = "pixel"
            shte.cell(1, 1).font = ffont
            shte.column_dimensions['A'].width = 13.57
            shte.row_dimensions[1].height = 80.0
            for yt in range(max(siz_reference[1], siz_sample[1])):
                shte.cell(yt+2, 1).value = yt
                shte.cell(yt+2, 1).font = ffont
            img_sheet = sheet_name[sh]
            file_count[img_sheet]
            for yt in range(siz_reference[1]):
                avreference = 0
                for xt in range(file_count[img_sheet]):
                    vreference = shte.cell(yt+2, 4+2*xt).value
                    #print(vreference)
                    avreference += vreference
                #print(float(avreference), float(file_count[img_sheet]))
                avreference = float(avreference)/float(file_count[img_sheet])
                shte.cell(yt+2, all+3).value = (avreference)
                shte.cell(yt+2, all+3).font = ffont
            for yt in range(siz_sample[1]):
                avsample =0
                for xt in range(file_count[img_sheet]):
                    vsample = shte.cell(yt+2, 3+2*xt).value
                    #print(vsample)
                    avsample += vsample
                #print(float(avsample), float(file_count[img_sheet]))
                avsample = float(avsample) / float(file_count[img_sheet])
                shte.cell(yt+2, all+2).value = (avsample)
                shte.cell(yt+2, all+2).font = ffont







        del wb["Sheet"]
        #刪除預設活頁簿
        os.chdir(targetpath)
        #print(targetpath)
        #更新資料夾位置到程式資料夾
        excelname = folderspath[folderspath.rfind('/')+1:]
        #自動抓取檔案名稱
        wb.save(excelname+'.xlsx')
        #保存活頁簿
        self.ui.start.setText("Start")
        #設定按鈕顯示名稱為START
        self.ui.start.setDisabled(False)
        #分析時START不可用
        QApplication.processEvents()
        #更新畫面
        

    


if __name__ == '__main__': #Python通用main函式
    import sys
    #系統模組
    app = QtWidgets.QApplication(sys.argv)
    #呼叫物件
    window = MainWindow_controller()
    #變數指向主要視窗
    window.show()
    #顯示視窗
    sys.exit(app.exec_())
    #按X的時候結束程式